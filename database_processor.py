import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Выгрузка:
file_path_1 = ""
# Отчет за предыдущий день:
file_path_2 = ""

# Путь к сохраненному файлу:
cleaned_file_path = ""


def download_and_parse(url):
    response = requests.get(url)
    response.raise_for_status()
    soup = BeautifulSoup(response.content, 'html.parser')

    training_programs = {}

    h3_elements = soup.find_all('h3')

    for h3 in h3_elements:
        training_direction = h3.get_text(strip=True)
        next_sibling = h3.find_next_sibling()
        while next_sibling and next_sibling.name != 'h3':
            if next_sibling.name == 'table':
                rows = next_sibling.find_all('tr')
                for row in rows:
                    columns = row.find_all('td')
                    if len(columns) == 2:
                        person_name = columns[1].get_text(strip=True)
                        if person_name in training_programs.keys():
                            training_programs[person_name].append(training_direction)
                        else:
                            training_programs[person_name] = [training_direction]
            next_sibling = next_sibling.find_next_sibling()

    return training_programs


urls = [
    "https://cpk.msu.ru/daily/dep_05_bs",
    "https://cpk.msu.ru/daily/dep_14_bs",
    "https://cpk.msu.ru/daily/dep_08_bs",
    "https://cpk.msu.ru/daily/dep_11_bs",
    "https://cpk.msu.ru/daily/dep_17_bs",
    "https://cpk.msu.ru/daily/dep_19_bs",
    "https://cpk.msu.ru/daily/dep_25_bs",
    "https://cpk.msu.ru/daily/dep_27_bs",
    "https://cpk.msu.ru/daily/dep_28_bs",
    "https://cpk.msu.ru/daily/dep_31_bs",
    "https://cpk.msu.ru/daily/dep_32_bs",
    "https://cpk.msu.ru/daily/dep_33_bs",
    "https://cpk.msu.ru/daily/dep_34_bs",
    "https://cpk.msu.ru/daily/dep_88_bs",
    "https://cpk.msu.ru/daily/dep_90_bs"
]

all_training_programs = {}

for url in urls:
    training_programs = download_and_parse(url)
    for person, programs in training_programs.items():
        if person in all_training_programs:
            all_training_programs[person].extend(programs)
        else:
            all_training_programs[person] = programs

msu_info = dict()

for person, programs in all_training_programs.items():
    msu_info[person.lower().replace('ё', 'е').strip()] = list(set(list(map(lambda x: x.replace('Направление подготовки ', '').replace('"', ''), programs))))


new_columns = [
    'КАТЕГОРИЯ АБИТУРИЕНТА',
    'ОТВЕТСТВЕННЫЙ',
    'КОД ЗВОНКА А',
    'КОММЕНТАРИЙ А',
    'КОД ЗВОНКА В',
    'КОММЕНТАРИЙ В',
    'КОД ЗВОНКА С',
    'КОММЕНТАРИЙ С'
]

df1 = pd.read_excel(file_path_1)
df1 = df1.dropna(axis=1, how='all')
df1 = df1.dropna(axis=0, how='all')
df1 = df1.iloc[10:].reset_index(drop=True)

df1_file_path = "C:/Users/fffkn/Downloads/newmetadf1.xlsx"
df1.to_excel(df1_file_path, index=False)
wb = openpyxl.load_workbook(df1_file_path)
ws = wb.active

ws.delete_rows(1)
wb.save(df1_file_path)

df1 = pd.read_excel(df1_file_path)

for col in new_columns:
    if col not in df1.columns:
        df1[col] = ''

col_index = df1.columns.get_loc("Список других ОП, на которые поданы документы")

df1.insert(col_index + 1, 'Внешние программы', '')


if file_path_2:
    df2 = pd.read_excel(file_path_2)
    df2 = df2.dropna(axis=1, how='all')
    df2 = df2.dropna(axis=0, how='all')

    reg_number_col_1 = [col for col in df1.columns if "регистрационный номер" in str(col).lower()][0]
    reg_number_col_2 = [col for col in df2.columns if "регистрационный номер" in str(col).lower()][0]

    df2.set_index(reg_number_col_2, inplace=True)

    for index, row in df1.iterrows():
        reg_number = row[reg_number_col_1]
        if reg_number in df2.index:
            for col in new_columns:
                if col in df2.columns:
                    # df1.at[index, col] = df2.at[reg_number, col]
                    results = df2.loc[reg_number]
                    if isinstance(results, pd.DataFrame):
                        for idx, row in results.iterrows():
                            val = row[col]
                            if not 'NaN' in str(val) and len(str(val)) > 0 and not pd.isna(val) and not pd.isnull(val):
                                df1.at[index, col] = val
                    elif isinstance(results, pd.Series):
                        val = results[col]
                        if not 'NaN' in str(val) and len(str(val)) > 0 and not pd.isna(val) and not pd.isnull(val):
                            df1.at[index, col] = val

categories = {'Поступление на места в рамках отдельной квоты': "Отдельная квота", 
              'Социальный лифт': 'Социальный лифт', '''Поступление на места в рамках квоты 
для лиц, имеющих особое право''': "Особая квота", 'Право поступления без вступительных испытаний': "БВИ",  '''Поступление на места в рамках квоты
целевого приема''': "Целевой прием"}

program = ""

if file_path_1.lower().__contains__('гму'):
    program = "gmu"
    minimum_value = 100
    discount_25_value = 100
    discount_50_value = 100
    discount_70_value = 100
    green_wave_value = 100

    responsibles = {"БВИ": "", "Особая квота": "", "Целевой прием": "",
                    "Отдельная квота": "", 'Социальный лифт': ""}
    
    paid_responsible = ""


elif file_path_1.lower().__contains__('соц'):
    program = "soc"
    minimum_value = 100
    discount_25_value = 100
    discount_50_value = 100
    discount_70_value = 100
    green_wave_value = 100

    responsibles = {"БВИ": "", "Особая квота": "", "Целевой прием": "",
                    "Отдельная квота": "", 'Социальный лифт': ""}
    paid_responsible = ""

elif file_path_1.lower().__contains__('полит'):
    program = "polit"
    minimum_value = 100
    discount_25_value = 100
    discount_50_value = 100
    discount_70_value = 100
    green_wave_value = 100

    responsibles = {"БВИ": "", "Особая квота": "", "Целевой прием": "",
                    "Отдельная квота": "", 'Социальный лифт': ""}
    paid_responsible = ""

elif file_path_1.lower().__contains__('пси'):
    program = "psych"
    minimum_value = 100
    discount_25_value = 100
    discount_50_value = 100
    discount_70_value = 100
    green_wave_value = 100

    responsibles = {"БВИ": "", "Особая квота": "", "Целевой прием": "",
                    "Отдельная квота": "", 'Социальный лифт': ""}
    paid_responsible = ""

else:
    raise ValueError

snils_info = dict()

with open('msses_sociology.txt', 'r') as file:
    content = file.read()
    numbers = content.split()
    for snils in numbers:
        if snils in snils_info.keys():
            snils_info[snils].append('Шанинка Социология')
        else:
            snils_info[snils] = ['Шанинка Социология']

with open('ranepa_sociology.txt', 'r') as file:
    content = file.read()
    numbers = content.split()
    for snils in numbers:
        if snils in snils_info.keys():
            snils_info[snils].append('РАНХИГС Социология')
        else:
            snils_info[snils] = ['РАНХИГС Социология']

with open('rudn_sociology.txt', 'r') as file:
    content = file.read()
    numbers = content.split()
    for snils in numbers:
        if snils in snils_info.keys():
            snils_info[snils].append('РУДН Социология')
        else:
            snils_info[snils] = ['РУДН Социология']

fio_info = dict()

df_fio = pd.read_csv('')
rsuh = df_fio.iloc[:, 1].tolist()

for el in rsuh:
    if el in fio_info.keys():
        fio_info[el].append('РГГУ Социология')
    else:
        fio_info[el] = ['РГГУ Социология']

for index in df1.index:

    snils = df1.at[index, 'СНИЛС / Уникальный идентификатор'].strip().replace(' ', '').replace('-', '')
    
    if snils in snils_info.keys():
        df1.at[index, 'Внешние программы'] = '; '.join(list(set(snils_info[snils])))

    name = df1.at[index, 'Фамилия, имя, отчество'].lower().replace('ё', 'е').strip()
    if name in msu_info.keys():
      df1.at[index, 'МГУ, ДВИ'] = '; '.join(msu_info[name])
    
    if name in fio_info.keys():
        df1.at[index, 'Внешние программы'] = str(df1.at[index, 'Внешние программы']) + '; '.join(list(set(fio_info[name])))
    
    value = ""
    if df1.at[index, 'Сумма конкурсных баллов'] >= minimum_value:
        value = "Полная стоимость"

    values = []
    points_sum = df1.at[index, 'Баллы за вступительные испытания']
    if points_sum >= discount_25_value:
        value = "Скидка 25%"
        if points_sum >= discount_50_value:
            value = "Скидка 50%"
            if points_sum >= discount_70_value:
                value = "Скидка 70%"

    if df1.at[index, 'Сумма конкурсных баллов'] >= green_wave_value:
        value = "Условно ЗВ"

    for category in categories.keys():
        if str(df1.at[index, category]).lower().__contains__('да'):
            values.append(categories[category])

    responsible = ""

    if values:
        responsible = responsibles[values[0]]
        if len(values) > 1:
            value = ' / '.join(values)
        else:
            value = values[0]
    else:
        if value != "":
            responsible = paid_responsible

        if program == "gmu":
            if value == "Условно ЗВ" and (
                    pd.isna(df1.at[index, "Приоритет иных мест"]) or int(df1.at[index, "Приоритет иных мест"]) < 3):
                responsible = ""

        elif program == "soc":
            if value == "Условно ЗВ":
                if (pd.isna(df1.at[index, "Приоритет иных мест"]) or int(df1.at[index, "Приоритет иных мест"]) < 3):
                    responsible = ""
                else:
                    responsible = ""

        elif program == "polit":
            if value == "Условно ЗВ" and (
                    pd.isna(df1.at[index, "Приоритет иных мест"]) or int(df1.at[index, "Приоритет иных мест"]) < 3):
                responsible = ""

        elif program == "psych":
            if value == "Условно ЗВ" and (
                    pd.isna(df1.at[index, "Приоритет иных мест"]) or int(df1.at[index, "Приоритет иных мест"]) < 3):
                responsible = ""
        else:
            raise ValueError

    df1.at[index, "КАТЕГОРИЯ АБИТУРИЕНТА"] = value
    df1.at[index, "ОТВЕТСТВЕННЫЙ"] = responsible

new_row = pd.DataFrame([df1.columns.tolist()], columns=df1.columns)
df1 = pd.concat([new_row, df1]).reset_index(drop=True)
df1.to_excel(cleaned_file_path, index=False)

wb = openpyxl.load_workbook(cleaned_file_path)
ws = wb.active

ws.delete_rows(1)

font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', size=10, bold=True)
alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin', color='000000'),
                     right=Side(style='thin', color='000000'),
                     top=Side(style='thin', color='000000'),
                     bottom=Side(style='thin', color='000000'))

target_column = None
for col in ws.iter_cols(min_row=0, max_row=1):
    for cell in col:
        if str(cell.value).lower() == "дата приказа о зачислении":
            target_column = cell.column
            break
    if target_column:
        break


column_width = (len("Всероссийская") + 2) * 1.2

gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
for cell in ws[1]:
    cell.fill = gray_fill

for row in ws.iter_rows():
    for cell in row:
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border

for col in ws.columns:
    column = col[0].column_letter
    ws.column_dimensions[column].width = column_width

highlight_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

keywords = ['Математика', 'Обществознание', 'Русский']

columns_to_highlight = []
for col in ws.iter_cols(min_row=1, max_row=1):
    for cell in col:
        if any(keyword in str(cell.value) for keyword in keywords):
            columns_to_highlight.append(cell.column_letter)

for col in columns_to_highlight:
    for cell in ws[col]:
        if cell.value in [100, '100']:
            cell.fill = highlight_fill

for col in ws.iter_cols(min_row=0, max_row=1):
    for cell in col:
        if "комментарий" in str(cell.value).lower():
            ws.column_dimensions[cell.column_letter].width = column_width * 3

if target_column:
    start_col_idx = target_column + 1
    for i, new_col_name in enumerate(new_columns):
        cell = ws.cell(row=1, column=start_col_idx + i)
        cell.value = new_col_name
        cell.font = bold_font
        cell.alignment = alignment
        cell.border = thin_border
        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

target_column = None
for col in ws.iter_cols(min_row=0, max_row=1):
    for cell in col:
        if str(cell.value).lower() == "внешние программы":
            target_column = cell.column
            break
    if target_column:
        break

new_columns = ['Внешние программы', "МГУ, ДВИ"]

if target_column:
    start_col_idx = target_column
    for i, new_col_name in enumerate(new_columns):
        cell = ws.cell(row=1, column=start_col_idx + i)
        cell.value = new_col_name
        cell.font = bold_font
        cell.alignment = alignment
        cell.border = thin_border
        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

wb.save(cleaned_file_path)